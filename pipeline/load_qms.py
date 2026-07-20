"""QMS load: clause-aware standard ingestion + structured KPI/NCR load.

Run:  py pipeline/load_qms.py        (after schema.py; regenerate via generate_qms.py)

Three stores, per the QMS architecture:
  - QS-001.pdf splits on CLAUSE NUMBERS, not token windows — a clause is the
    natural retrieval unit; each becomes a :Clause node with an embedding
  - quality_kpis.xlsx -> :QualityKPI typed rows. Numbers are NEVER embedded;
    the analytics leg of the Ask agent computes over them with Cypher
  - ncr_register.xlsx -> :NCR nodes linking equipment -> clause, which is what
    lets 'why is PPM rising' traverse to the governing clause

Clause APPLIES_TO equipment resolves from class phrases in the clause text
(same philosophy as tag mentions in SOPs — regex first, no LLM on clean data).
"""

import json
import re
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

from neo4j import GraphDatabase
from openpyxl import load_workbook
from pypdf import PdfReader

sys.path.insert(0, str(Path(__file__).parent))
from config import (NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD, NEO4J_DATABASE,
                    CORPUS, OLLAMA_BASE_URL, EMBED_MODEL, EMBED_DIM)

CLAUSE_RE = re.compile(r"^(\d+(?:\.\d+)+|\d+)\s+([A-Z][^\n]{3,80})$")
CLASS_HINTS = {
    "centrifugal_pump": ["pump", "fluid transfer", "strainer", "seal integrity"],
    "heat_exchanger": ["heat exchanger", "exchanger"],
    "control_valve": ["control valve", "valve positioner"],
}
STD_ID = "QS-001"


def ollama_embed(texts):
    req = urllib.request.Request(
        f"{OLLAMA_BASE_URL}/api/embed",
        data=json.dumps({"model": EMBED_MODEL, "input": texts}).encode(),
        headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=300) as r:
        vecs = json.loads(r.read())["embeddings"]
    assert all(len(v) == EMBED_DIM for v in vecs), "embedding dimension drift!"
    return vecs


def parse_clauses(pdf_path):
    """Clause-aware split: new unit at every clause-number heading."""
    reader = PdfReader(str(pdf_path))
    lines = []
    for pno, page in enumerate(reader.pages, 1):
        for ln in (page.extract_text() or "").splitlines():
            lines.append((pno, ln.strip()))
    clauses, current = [], None
    for pno, ln in lines:
        m = CLAUSE_RE.match(ln)
        if m:
            if current:
                clauses.append(current)
            current = {"clause_id": m.group(1), "title": m.group(2).strip(),
                       "page": pno, "text": ""}
        elif current is not None and ln:
            current["text"] += (" " if current["text"] else "") + ln
    if current:
        clauses.append(current)
    return [c for c in clauses if c["text"]]


def rows(path):
    ws = load_workbook(path, read_only=True).active
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(it)]
    for r in it:
        if any(v is not None for v in r):
            yield dict(zip(header, r))


def main():
    pdf = CORPUS / "qms" / "QS-001.pdf"
    clauses = parse_clauses(pdf)
    vecs = ollama_embed([f"{c['clause_id']} {c['title']}\n{c['text']}"
                         for c in clauses])
    for c, v in zip(clauses, vecs):
        c["embedding"] = v
        low = c["text"].lower()
        c["classes"] = [k for k, hints in CLASS_HINTS.items()
                        if any(h in low for h in hints)]

    kpis = [{"process_id": r["process_id"], "tag": r["equipment_tag"],
             "desc": r.get("process_desc", ""), "metric": r["metric"],
             "period": str(r["period"]), "value": float(r["value"])}
            for r in rows(CORPUS / "quality_kpis.xlsx")]

    ncrs = []
    for r in rows(CORPUS / "ncr_register.xlsx"):
        d = r["date"]
        ncrs.append({"ncr_id": r["ncr_id"],
                     "date": (d.date().isoformat() if isinstance(d, datetime)
                              else str(d)),
                     "tag": r["equipment_tag"], "clause_id": str(r["clause_id"]),
                     "description": r["description"], "status": r["status"]})

    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD),
                                  notifications_min_severity="OFF")
    with driver.session(database=NEO4J_DATABASE) as s:
        s.run("MERGE (st:Standard {std_id: $id}) "
              "SET st.title = 'Plant Site A Quality Management Standard', "
              "st.overlay = 'automotive customer overlay (IATF 16949 aligned)'",
              id=STD_ID)
        s.run("""
            UNWIND $clauses AS c
            MATCH (st:Standard {std_id: $std})
            MERGE (cl:Clause {clause_id: c.clause_id})
            SET cl.title = c.title, cl.text = c.text, cl.page = c.page,
                cl.embedding = c.embedding, cl.std_id = $std
            MERGE (st)-[:HAS_CLAUSE]->(cl)
            WITH cl, c
            UNWIND c.classes AS klass
            MATCH (e:Equipment {iso14224_class: klass})
            MERGE (cl)-[:APPLIES_TO]->(e)
        """, clauses=[{k: c[k] for k in
                       ("clause_id", "title", "text", "page", "embedding",
                        "classes")} for c in clauses], std=STD_ID)
        s.run("""
            UNWIND $kpis AS k
            MERGE (p:Process {process_id: k.process_id})
            SET p.description = k.desc, p.tag = k.tag
            WITH p, k
            MATCH (e:Equipment {tag: k.tag})
            MERGE (p)-[:USES]->(e)
            MERGE (q:QualityKPI {process_id: k.process_id, metric: k.metric,
                                 period: k.period})
            SET q.value = k.value, q.tag = k.tag
            MERGE (q)-[:OF]->(p)
            MERGE (q)-[:FOR]->(e)
        """, kpis=kpis)
        s.run("""
            UNWIND $ncrs AS n
            MERGE (x:NCR {ncr_id: n.ncr_id})
            SET x.date = date(n.date), x.description = n.description,
                x.status = n.status
            WITH x, n
            MATCH (e:Equipment {tag: n.tag})
            MERGE (x)-[:AGAINST]->(e)
            WITH x, n
            MATCH (cl:Clause {clause_id: n.clause_id})
            MERGE (x)-[:CITES]->(cl)
        """, ncrs=ncrs)
        counts = s.run("""
            UNWIND ['Standard','Clause','QualityKPI','NCR'] AS lbl
            CALL (lbl) { MATCH (n) WHERE lbl IN labels(n) RETURN count(n) AS c }
            RETURN lbl, c
        """).data()
        applies = s.run("MATCH (:Clause)-[r:APPLIES_TO]->() RETURN count(r) AS n"
                        ).single()["n"]
    driver.close()
    print(f"parsed {len(clauses)} clauses from {pdf.name}: "
          f"{[c['clause_id'] for c in clauses]}")
    print("loaded:", {c["lbl"]: c["c"] for c in counts},
          f"+ {applies} clause APPLIES_TO links")


if __name__ == "__main__":
    main()
